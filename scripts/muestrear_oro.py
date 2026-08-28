"""Muestrea las 150 preguntas de la muestra de oro y arma la hoja de codificación.

La muestra de oro no es el dataset: es la **calibración**. Sirve para medir si
las ~22 mil que clasificó el modelo son confiables. Por eso se codifica antes de
que el modelo corra sobre el corpus, y a ciegas.

«A ciegas» son tres cosas, y las tres se implementan aquí:

1. El humano no ve la salida del modelo. Esta hoja no la incluye.
2. La hoja **no lleva nombre de periodista ni medio**. Ni siquiera lleva el
   `id_pregunta`, que trae la fecha y el número de hilo y permitiría buscar de
   quién es. La hoja usa un código opaco (`P-001`) y la llave de unión vive en
   un archivo aparte que no hay que abrir mientras se codifica.
3. El texto va **redactado**: el 10% de las preguntas dice el medio dentro de su
   propio turno, y el contexto puede traer a la presidenta llamando al
   periodista por su nombre («Gracias, Hans»). Ambas cosas se tapan, y al final
   se **verifica fila por fila** que ningún nombre de la llave sobrevive en el
   texto visible. Las filas que no pasan se descartan y se sortea reemplazo.

    python scripts/muestrear_oro.py [--n 150] [--lote1 30] [--semilla 20260828]
"""

from __future__ import annotations

import argparse
import csv
import json
import random
import re
import sys
import unicodedata
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from estenograficas.config import paths  # noqa: E402
from estenograficas.parser import redactar_identificacion  # noqa: E402

MARCA = "[identificación removida]"
MARCA_NOMBRE = "[nombre removido]"

# Preguntas que el humano ya vio en las dos hojas de ejemplo. Quedan fuera del
# marco muestral: ya las codificó sabiendo que eran ejemplos, y la v1 además se
# generó antes de implementar la redacción, así que mostraba el medio.
YA_VISTAS = {
    "2026-03-25-h0-t88", "2026-01-23-h5-t125", "2024-12-20-h5-t110",
    "2026-06-25-h3-t91", "2026-06-09-h2-t88", "2025-11-05-h2-t52",
    "2026-07-13-h1-t66", "2025-08-21-h0-t47", "2026-02-06-h6-t105",
    "2024-12-05-h3-t111", "2026-05-25-h0-t124", "2026-04-29-h0-t17",
    "2026-01-13-h2-t130", "2025-03-25-h0-t28", "2025-11-18-h4-t128",
    "2025-12-22-h1-t86", "2025-05-15-h5-t96", "2026-02-27-h8-t180",
    "2026-03-20-h2-t56", "2025-12-15-h2-t82", "2024-12-17-h1-t26",
    "2025-06-13-h7-t140", "2026-08-19-h1-t69", "2025-03-11-h2-t46",
    "2025-10-24-h0-t54", "2024-10-31-h2-t41",
}

# Una sola dimensión, cuatro valores. Decidido el 2026-08-28.
#
# El libro de códigos original tenía cuatro dimensiones (objetivo, postura,
# funcion, insistencia). Se colapsan a una porque el objetivo del trabajo dejó
# de ser un artículo académico: la pregunta es de qué signo son las preguntas,
# y `funcion` e `insistencia` no la contestan. Codificar una columna en vez de
# cuatro baja el trabajo humano de 600 decisiones a 150.
#
# Lo que NO se hizo, y es la parte que importa: colapsar a las tres categorías
# de uso periodístico —crítica / afín / de interés público—. «De interés
# público» no está en el mismo eje que las otras dos: mide el mérito de la
# pregunta, no su dirección, y casi todo el buen periodismo es crítico y de
# interés público a la vez. Obligar a escoger hace que quien codifica resuelva
# ese conflicto en su cabeza, distinto cada vez, y eso hunde el alfa. Además
# la etiqueta afirma que las otras dos no son de interés público, que es
# justo el tipo de juicio que la regla 8 prohíbe.
#
# `critica_a_un_tercero` existe porque el 15% de las preguntas habla de la
# oposición o de un actor externo. Una pregunta durísima contra García Luna o
# contra Trump no es crítica al gobierno ni lo halaga, y sin esta categoría
# caería en el cajón neutral junto con las peticiones de dato.
OPCIONES = {
    "postura": ["crítica al gobierno", "afín al gobierno", "crítica a un tercero",
                "neutral", "no clasificable"],
}


def _sin_acentos(s: str) -> str:
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if not unicodedata.combining(c))


def tapa_nombres(texto: str, periodista: str | None, medio: str | None) -> str:
    """Tapa el nombre del periodista y su medio dondequiera que aparezcan.

    Se tapa el nombre completo siempre. De los tokens sueltos solo se tapan los
    **apellidos** (del segundo token en adelante) y los nombres de pila cuando
    vienen como vocativo —«Gracias, Hans»—, porque tapar `Manuel` a secas
    destroza `Andrés Manuel López Obrador`, que aparece en medio corpus y no es
    una fuga.
    """
    for entero in filter(None, (periodista, medio)):
        texto = re.sub(re.escape(entero), MARCA_NOMBRE, texto, flags=re.IGNORECASE)
    if periodista:
        partes = periodista.split()
        for apellido in partes[1:]:
            if len(apellido) > 3:
                texto = re.sub(rf"\b{re.escape(apellido)}\b", MARCA_NOMBRE, texto)
        if partes and len(partes[0]) > 2:
            texto = re.sub(
                rf"(?<=[,:;¡!¿?]\s){re.escape(partes[0])}\b", MARCA_NOMBRE, texto
            )
    return texto


# Palabras que aparecen en los nombres de medio pero no identifican a nadie.
# Sin esta lista, «México» cuenta como fuga y se descarta toda pregunta que lo
# diga, que en este corpus es un sesgo grande y en una sola dirección: se irían
# justo las preguntas de tema nacional.
_GENERICAS = {
    "grupo", "media", "medios", "noticias", "noticiero", "noticiera", "radio",
    "television", "televisora", "revista", "periodico", "diario", "red", "redes",
    "canal", "prensa", "informativo", "informativa", "programa", "mexico",
    "nacional", "digital", "portal", "agencia", "comunicacion", "comunicaciones",
    "telecomunicaciones", "editorial", "grupo", "news", "press", "online",
    "luces", "网", "internacional", "libre", "publica", "publico", "the",
}


def hay_fuga(visible: str, periodista: str | None, medio: str | None) -> str | None:
    """Devuelve el término filtrado, o None. Es la verificación, no una promesa.

    Solo cuenta como fuga lo que **identifica**: el nombre completo, el nombre
    del medio completo, los apellidos y los tokens distintivos del medio
    (`Contralínea`, `Milenio`). Una palabra genérica del nombre del medio no
    revela nada y descartar por ella sesga la muestra.
    """
    plano = _sin_acentos(visible).lower()
    terminos = []
    if periodista:
        terminos += [periodista] + [x for x in periodista.split()[1:] if len(x) > 3]
    if medio:
        terminos.append(medio)
        terminos += [
            x for x in re.findall(r"[\wÁÉÍÓÚÑÜáéíóúñü]+", medio)
            if len(x) > 4 and _sin_acentos(x).lower() not in _GENERICAS
        ]
    for t in terminos:
        if re.search(rf"\b{re.escape(_sin_acentos(t).lower())}\b", plano):
            return t
    return None


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--n", type=int, default=150)
    ap.add_argument("--lote1", type=int, default=30)
    ap.add_argument("--semilla", type=int, default=20260828)
    args = ap.parse_args()

    p = paths()
    universo = []
    for linea in p.hilos.read_text(encoding="utf-8").splitlines():
        if not linea.strip():
            continue
        h = json.loads(linea)
        turnos = h["turnos"]
        etiqueta = lambda t: ("PRENSA: " if t["rol"] == "pregunta" else "PRESIDENTA: ")
        for i, t in enumerate(turnos):
            if t["rol"] != "pregunta" or t["ruido"]:
                continue
            pid = f'{h["conferencia_id"]}-h{h["hilo"]}-t{t["orden"]}'
            if pid in YA_VISTAS or len(t["texto"].strip()) < 12:
                continue
            # Contexto previo: hasta 4 turnos, y se corta al llegar a ~900
            # caracteres. Dos turnos fijos no alcanzaban: cuando la pregunta
            # viene interrumpida en la estenográfica —pasa, y la marca es que
            # termina en puntos suspensivos— el turno anterior puede ser de
            # otro tema por completo y el fragmento queda ilegible.
            previos, largo = [], 0
            for u in reversed(turnos[max(0, i - 4):i]):
                s = etiqueta(u) + " ".join(u["texto"].split())
                if not s.strip() or s.strip().endswith(":"):
                    continue
                if previos and largo + len(s) > 900:
                    break
                previos.append(s)
                largo += len(s)
            previos.reverse()
            # Lo que siguió. Va rotulado aparte para que se lea como contexto y
            # no como parte de la pregunta.
            #
            # Cuando la pregunta viene **interrumpida** en la estenográfica, un
            # solo turno no basta: el sentido suele aparecer dos o tres turnos
            # después, cuando el periodista termina la idea. Medido: 16 de las
            # 150 quedan cortadas. Para ésas se muestra el intercambio completo,
            # hasta 4 turnos.
            truncada = t["texto"].rstrip().endswith(("…", "...", "—", "–"))
            sig, largo_sig = [], 0
            for u in turnos[i + 1:i + (5 if truncada else 3)]:
                s = " ".join(u["texto"].split())
                if not s:
                    continue
                sig.append(etiqueta(u) + s)
                largo_sig += len(s)
                if not truncada or largo_sig > 600:
                    break
            sigue = "\n".join(sig)
            universo.append({
                "id": pid,
                "fecha": h["conferencia_id"],
                "periodista": h["periodista"],
                "medio": h["medio"],
                "texto": t["texto"],
                "contexto": previos,
                "sigue": sigue,
                "truncada": truncada,
                "abre_hilo": t["atribucion"] == "declarada",
            })

    print(f"marco muestral: {len(universo):,} preguntas "
          f"({len(YA_VISTAS)} excluidas por ya vistas)")

    rnd = random.Random(args.semilla)
    barajado = universo[:]
    rnd.shuffle(barajado)

    elegidas, descartadas = [], []
    for r in barajado:
        if len(elegidas) == args.n:
            break
        texto, _ = redactar_identificacion(r["texto"])
        texto = tapa_nombres(texto, r["periodista"], r["medio"])
        tapa = lambda s: tapa_nombres(redactar_identificacion(s)[0],
                                      r["periodista"], r["medio"])
        ctx = "\n".join(tapa(c) for c in r["contexto"])
        sigue = tapa(r["sigue"]) if r["sigue"] else ""
        fuga = hay_fuga(" ".join((texto, ctx, sigue)), r["periodista"], r["medio"])
        if fuga:
            descartadas.append((r["id"], fuga))
            continue
        elegidas.append({**r, "visible": " ".join(texto.split()), "ctx": ctx,
                         "sig": sigue})

    if len(elegidas) < args.n:
        print(f"ATENCIÓN: solo se juntaron {len(elegidas)} de {args.n}", file=sys.stderr)

    orden = list(range(len(elegidas)))
    rnd.shuffle(orden)
    lote = {i: (1 if k < args.lote1 else 2) for k, i in enumerate(orden)}

    p.gold.mkdir(parents=True, exist_ok=True)
    llave = p.gold / "muestra_oro_LLAVE_no_abrir.csv"
    with llave.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["codigo", "lote", "id_pregunta", "fecha", "periodista", "medio",
                    "abre_hilo", "truncada"])
        for i, r in enumerate(elegidas, 1):
            w.writerow([f"P-{i:03d}", lote[i - 1], r["id"], r["fecha"],
                        r["periodista"] or "", r["medio"] or "", r["abre_hilo"],
                        r["truncada"]])

    try:
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        print("openpyxl no está; solo se escribió la llave", file=sys.stderr)
        return 1

    # Si ya existe una hoja con respuestas, se arrastran por código. Los códigos
    # son estables mientras no cambie la semilla, así que regenerar la hoja por
    # un arreglo de formato no puede costarle al humano el trabajo ya hecho.
    previas: dict[str, list] = {}
    anterior = p.gold / "muestra_oro_hoja.xlsx"
    if anterior.exists():
        try:
            from openpyxl import load_workbook
            wba = load_workbook(anterior, data_only=True)
            for ws in wba.worksheets:
                cab = [c.value for c in ws[1]]
                for fila in ws.iter_rows(min_row=2, values_only=True):
                    d = dict(zip(cab, fila))
                    vals = [d.get("postura"),
                            d.get("fragmento que te hizo decidir"),
                            d.get("notas / dudas")]
                    if d.get("codigo") and any(v for v in vals):
                        previas[d["codigo"]] = vals
            if previas:
                print(f"se arrastran {len(previas)} respuestas ya codificadas")
        except Exception as exc:  # openpyxl no pudo abrirla; no es fatal
            print(f"no se pudo leer la hoja anterior ({exc})", file=sys.stderr)

    wb = Workbook()
    for numero in (1, 2):
        ws = wb.active if numero == 1 else wb.create_sheet()
        ws.title = f"lote {numero}"
        ws.append(["codigo", "lo que se dijo antes", "PREGUNTA A CODIFICAR",
                   "lo que siguió", "postura", "fragmento que te hizo decidir",
                   "notas / dudas"])
        filas = [(i, r) for i, r in enumerate(elegidas) if lote[i] == numero]
        for i, r in filas:
            visible = r["visible"] + ("   ⟨la pregunta queda cortada aquí en la "
                                      "versión estenográfica⟩" if r["truncada"] else "")
            cod = f"P-{i + 1:03d}"
            ws.append([cod, r["ctx"], visible, r["sig"], *previas.get(cod, ["", "", ""])])
        dv = DataValidation(
            type="list", formula1='"' + ",".join(OPCIONES["postura"]) + '"',
            allow_blank=True, showDropDown=False,
        )
        ws.add_data_validation(dv)
        dv.add(f"E2:E{len(filas) + 1}")
        for col, ancho in zip("ABCDEFG", (9, 50, 68, 44, 21, 36, 30)):
            ws.column_dimensions[col].width = ancho
        ws.freeze_panes = "E2"
        for fila in ws.iter_rows(min_row=1, max_row=len(filas) + 1):
            for celda in fila:
                celda.alignment = celda.alignment.copy(wrap_text=True, vertical="top")

    hoja = p.gold / "muestra_oro_hoja.xlsx"
    try:
        wb.save(hoja)
    except PermissionError:
        # Suele significar que está abierta en Excel. No se pisa y no se pierde
        # nada: se escribe al lado con sufijo y el humano decide.
        n = 2
        while (p.gold / f"muestra_oro_hoja_v{n}.xlsx").exists():
            n += 1
        hoja = p.gold / f"muestra_oro_hoja_v{n}.xlsx"
        wb.save(hoja)
        print(f"la hoja anterior estaba abierta; se escribió {hoja.name}",
              file=sys.stderr)

    n1 = sum(1 for v in lote.values() if v == 1)
    print(f"\nhoja  : {hoja}   (lote 1: {n1} · lote 2: {len(elegidas) - n1})")
    print(f"llave : {llave}   NO ABRIR mientras se codifica")
    print(f"descartadas por fuga de nombre: {len(descartadas)}")
    for pid, term in descartadas[:6]:
        print(f"    {pid}  filtraba «{term}»")

    from collections import Counter
    print("\ncomposición de las 150:")
    print("  por año  :", dict(Counter(r["fecha"][:4] for r in elegidas)))
    print("  abren hilo:", sum(1 for r in elegidas if r["abre_hilo"]),
          "· son seguimiento:", sum(1 for r in elegidas if not r["abre_hilo"]))
    largos = sorted(len(r["visible"]) for r in elegidas)
    print(f"  largo    : mediana {largos[len(largos) // 2]} caracteres, "
          f"de {largos[0]} a {largos[-1]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
